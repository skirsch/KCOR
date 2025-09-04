import openpyxl
from openpyxl import load_workbook
import shutil

# First, make a copy of the original file
shutil.copy('KCORv4.xlsx', 'KCORv4_with_formulas.xlsx')

# Load the copied workbook
wb = load_workbook('KCORv4_with_formulas.xlsx')
ws = wb['2022-06 booster']

print("Filling in missing formulas in columns BH, BI, BJ...")

# Column BH (v3/v0) - similar to BG (v2/v0) but using O (v3) instead of N (v2)
bh_formulas = {
    17: "=OFFSET(O10,$M$1-1,0)",
    18: "=OFFSET(L10, $M$1-1,0)",
    19: "=BH11*BH17*BH20/BH18",
    20: "=OFFSET(L1, $BE$22-1,0)*(Z$6^($BE$22-10))",
    21: "=BH19*BH18/(BH17*BH20)",
    24: "=OFFSET(O$1, $BE$22-1,0)"
}

# Column BI (v3/v1) - similar to BF (v2/v1) but using O (v3) instead of N (v2)
bi_formulas = {
    17: "=OFFSET(O10,$M$1-1,0)",
    18: "=OFFSET(M10,$M$1-1,0)",
    19: "=BI11*BI17*BI20/BI18",
    20: "=OFFSET(M1, $BE$22-1,0)",
    21: "=BI19*BI18/(BI17*BI20)",
    24: "=OFFSET(O$1, $BE$22-1,0)"
}

# Column BJ (v3/v2) - similar to BF (v2/v1) but using O (v3) instead of N (v2), and N (v2) instead of M (v1)
bj_formulas = {
    17: "=OFFSET(O10,$M$1-1,0)",
    18: "=OFFSET(N10,$M$1-1,0)",
    19: "=BJ11*BJ17*BJ20/BJ18",
    20: "=OFFSET(N1, $BE$22-1,0)",
    21: "=BJ19*BJ18/(BJ17*BJ20)",
    24: "=OFFSET(O$1, $BE$22-1,0)"
}

# Apply the formulas
for row, formula in bh_formulas.items():
    ws.cell(row=row, column=60).value = formula  # BH = column 60
    print(f"BH{row}: {formula}")

for row, formula in bi_formulas.items():
    ws.cell(row=row, column=61).value = formula  # BI = column 61
    print(f"BI{row}: {formula}")

for row, formula in bj_formulas.items():
    ws.cell(row=row, column=62).value = formula  # BJ = column 62
    print(f"BJ{row}: {formula}")

# Save the workbook
wb.save('KCORv4_with_formulas.xlsx')
print(f"\nFormulas filled in and saved to KCORv4_with_formulas.xlsx")

wb.close()
