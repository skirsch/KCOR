import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('KCORv4.xlsx', data_only=False)
ws = wb['2022-06 booster']

print("Sheet dimensions:", ws.dimensions)
print("Max row:", ws.max_row)
print("Max column:", ws.max_column)

# Look at the 3 columns to the left of BH, BI, BJ for rows 17-24
print("\nColumns to the left of BH, BI, BJ (rows 17-24):")
for col in [57, 58, 59]:  # BE, BF, BG (3 columns to the left of BH)
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"\nColumn {col_letter}:")
    for row in range(17, 25):
        cell = ws.cell(row=row, column=col)
        cell_value = cell.value
        if cell_value is not None:
            print(f"Row {row}: Value={cell_value}")

# Look at the specific columns BH, BI, BJ for rows 17-24
print("\nColumns BH, BI, BJ (rows 17-24) - current values:")
for col in [60, 61, 62]:  # BH, BI, BJ
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"\nColumn {col_letter}:")
    for row in range(17, 25):
        cell = ws.cell(row=row, column=col)
        cell_value = cell.value
        if cell_value is not None:
            print(f"Row {row}: Value={cell_value}")

# Look at the ratio headings in row 10
print("\nRatio headings in row 10:")
for col in [57, 58, 59, 60, 61, 62]:  # BE, BF, BG, BH, BI, BJ
    col_letter = openpyxl.utils.get_column_letter(col)
    cell_value = ws.cell(row=10, column=col).value
    if cell_value is not None:
        print(f"Column {col_letter}: {cell_value}")

wb.close()
