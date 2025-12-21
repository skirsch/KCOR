"""
Generate figures for the KCOR methods paper: pathological frailty-mixture negative control.

This script reads parameters from the Excel workbook `example/Frail_cohort_mix.xlsx` (sheet: "1-10 frailty")
and produces two figures:

1) neg_control_pathological_fit.png
   - observed cohort hazard h(t) from the frailty-mixture construction
   - best-fit 4-parameter generalized relaxation model overlay

2) neg_control_pathological_kcor.png
   - pre-normalization hazard curves for two cohorts with different frailty-mixture weights (null effect)
   - post-normalization KCOR(t) after independent curvature fitting and early-window anchoring

Outputs are written to `documentation/preprint/figs/` by default.

Dependencies: numpy, scipy, matplotlib, openpyxl
"""

from __future__ import annotations

import argparse
import json
import math
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.optimize import least_squares


def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


@dataclass(frozen=True)
class RelaxFit:
    C: float
    k0: float
    k_inf: float
    tau: float
    rms_log_error: float


def relaxation_log_hazard(s: np.ndarray, C: float, k0: float, k_inf: float, tau: float) -> np.ndarray:
    """
    log h(s) = C + k_inf*s + (k0 - k_inf)*tau*(1 - exp(-s/tau))
    """
    s = np.asarray(s, dtype=float)
    tau = float(tau)
    return C + k_inf * s + (k0 - k_inf) * tau * (1.0 - np.exp(-s / (tau + 1e-12)))


def fit_relaxation_model(weeks: np.ndarray, hazard: np.ndarray) -> RelaxFit:
    """
    Fit the 4-parameter relaxation model to log hazard via bounded least squares.
    Parameterization enforces k_inf >= k0 (delta_k >= 0) and tau > 0 for stability.
    """
    weeks = np.asarray(weeks, dtype=float)
    hazard = np.asarray(hazard, dtype=float)

    mask = np.isfinite(weeks) & np.isfinite(hazard) & (hazard > 0)
    w = weeks[mask]
    h = hazard[mask]
    if len(w) < 10:
        raise ValueError(f"Not enough valid points to fit: {len(w)}")

    logh = np.log(h)

    # Initial guesses
    C0 = float(logh[0])
    # Rough early slope (can be negative)
    k0_0 = float((logh[min(5, len(logh) - 1)] - logh[0]) / (w[min(5, len(w) - 1)] - w[0] + 1e-12))
    # Rough late slope
    tail_n = min(30, len(logh) - 1)
    k_inf_0 = float((logh[-1] - logh[-tail_n]) / (w[-1] - w[-tail_n] + 1e-12))
    delta_k0 = max(k_inf_0 - k0_0, 1e-4)
    tau0 = max(1.0, float((w.max() - w.min()) / 4.0))

    # p = [C, k0, delta_k, tau], where k_inf = k0 + delta_k
    p0 = np.array([C0, k0_0, delta_k0, tau0], dtype=float)
    lower = np.array([-np.inf, -np.inf, 0.0, 1e-6], dtype=float)
    upper = np.array([np.inf, np.inf, np.inf, np.inf], dtype=float)

    def residuals(p: np.ndarray) -> np.ndarray:
        C, k0, delta_k, tau = p
        k_inf = k0 + delta_k
        pred = relaxation_log_hazard(w, C=C, k0=k0, k_inf=k_inf, tau=tau)
        return logh - pred

    res = least_squares(
        residuals,
        p0,
        bounds=(lower, upper),
        method="trf",
        max_nfev=10_000,
        ftol=1e-12,
        xtol=1e-12,
        gtol=1e-12,
    )
    C, k0, delta_k, tau = res.x
    k_inf = k0 + delta_k
    rms = float(np.sqrt(np.mean(residuals(res.x) ** 2)))
    return RelaxFit(C=float(C), k0=float(k0), k_inf=float(k_inf), tau=float(tau), rms_log_error=rms)


@dataclass(frozen=True)
class PathologicalSheetData:
    weekly_log_slope: float
    base_prob: float
    frailty: list[float]
    weeks: np.ndarray
    p_by_group: np.ndarray  # shape (n_weeks, n_groups) death probabilities per group per interval
    hazard_sheet: np.ndarray  # cohort hazard from sheet (may be used for verification)
    alive0_by_group: np.ndarray  # initial alive counts (from sheet)


def read_pathological_sheet(xlsx_path: Path, sheet_name: str) -> PathologicalSheetData:
    """
    Read the pathological frailty-mixture construction from the Excel sheet.

    Expected layout (as in example/Frail_cohort_mix.xlsx, sheet '1-10 frailty'):
    - B1: weekly log slope
    - B2: base hazard/probability for frailty=1
    - B3:F3: frailty values
    - Row 6..: week (A), per-group p(t) in B..F, per-group alive counts in H..L, cohort hazard in T
    """
    # IMPORTANT: Do NOT use read_only=True with random cell access; it becomes extremely slow
    # (each ws["A123"] style access can trigger streaming XML scans). This workbook is small enough
    # to load normally.
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    weekly_log_slope = float(ws["B1"].value)
    base_prob = float(ws["B2"].value)
    frailty = [float(ws[c].value) for c in ("B3", "C3", "D3", "E3", "F3")]

    weeks: list[float] = []
    p_rows: list[list[float]] = []
    hazard_sheet: list[float] = []
    alive0_by_group: np.ndarray | None = None

    # Data table starts at row 6 (week=0). Read sequentially for speed.
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=20, values_only=True):
        # Columns are: A..T => indices 0..19
        w = row[0]
        if w is None:
            break
        try:
            w = float(w)
        except Exception:
            break

        p = row[1:6]  # B..F
        if any(v is None for v in p):
            break
        p = [float(v) for v in p]

        h = row[19]  # T
        if h is None:
            break
        h = float(h)

        if alive0_by_group is None:
            alive0 = row[7:12]  # H..L
            if any(v is None for v in alive0):
                raise ValueError("Missing initial alive counts in H..L for the first data row.")
            alive0_by_group = np.array([float(v) for v in alive0], dtype=float)

        weeks.append(w)
        p_rows.append(p)
        hazard_sheet.append(h)

    if alive0_by_group is None:
        raise ValueError("Could not read initial alive counts (H..L) from the sheet.")

    return PathologicalSheetData(
        weekly_log_slope=weekly_log_slope,
        base_prob=base_prob,
        frailty=frailty,
        weeks=np.array(weeks, dtype=float),
        p_by_group=np.array(p_rows, dtype=float),
        hazard_sheet=np.array(hazard_sheet, dtype=float),
        alive0_by_group=alive0_by_group.astype(float),
    )


def simulate_cohort_hazard_from_p(
    p_by_group: np.ndarray,
    weights: np.ndarray,
    N0: float,
) -> np.ndarray:
    """
    Deterministic discrete-time cohort simulation given per-group death probabilities p_i(t).

    Mirrors the spreadsheet logic:
    - deaths_i(t) = alive_i(t) * p_i(t)
    - alive_i(t+1) = alive_i(t) - deaths_i(t)
    - cohort hazard: h(t) = -ln(1 - sum(deaths_i)/sum(alive_i))
    """
    p_by_group = np.asarray(p_by_group, dtype=float)
    weights = np.asarray(weights, dtype=float)
    weights = weights / weights.sum()

    n_steps, n_groups = p_by_group.shape
    alive = (N0 * weights).astype(float)
    hazards = np.zeros(n_steps, dtype=float)

    for t in range(n_steps):
        p = p_by_group[t, :]
        # Guard: if any p >= 1, clamp slightly below 1 to avoid negative survival
        p = np.clip(p, 0.0, 1.0 - 1e-12)
        deaths = alive * p
        tot_alive = float(alive.sum())
        tot_deaths = float(deaths.sum())
        if tot_alive <= 0:
            hazards[t] = np.nan
            continue
        frac = tot_deaths / tot_alive
        frac = min(max(frac, 0.0), 1.0 - 1e-12)
        hazards[t] = -math.log(1.0 - frac)
        alive = alive - deaths

    return hazards


def curvature_normalize(h: np.ndarray, weeks: np.ndarray, fit: RelaxFit) -> np.ndarray:
    """
    Remove the fitted shape term g(s) from hazard series h(s):
      log h(s) = C + g(s)  =>  h_adj(s) = h(s) * exp(-g(s))
    """
    weeks = np.asarray(weeks, dtype=float)
    h = np.asarray(h, dtype=float)
    g = fit.k_inf * weeks + (fit.k0 - fit.k_inf) * fit.tau * (1.0 - np.exp(-weeks / (fit.tau + 1e-12)))
    return h * np.exp(-g)


def compute_kcor(
    weeks: np.ndarray,
    hA: np.ndarray,
    hB: np.ndarray,
    skip_weeks: int,
    normalization_weeks: int,
    fitA: RelaxFit,
    fitB: RelaxFit,
) -> np.ndarray:
    """
    Compute KCOR(t) from two hazard series under generalized curvature normalization + baseline anchoring.

    This matches the repo's core KCOR normalization pattern:
    - Apply DYNAMIC_HVE_SKIP_WEEKS: do not accumulate hazard during the first `skip_weeks`.
    - Compute cumulative adjusted hazard CH for each cohort.
    - Form K_raw(t) = CH_A(t) / CH_B(t).
    - Normalize to 1 at an "effective baseline week" index:
        t0_idx = skip_weeks + normalization_weeks - 1
      so that KCOR(t) = K_raw(t) / K_raw(t0_idx).
    """
    hA_adj = curvature_normalize(hA, weeks, fitA)
    hB_adj = curvature_normalize(hB, weeks, fitB)

    # Apply skip-weeks (start accumulating after early transient)
    hA_eff = np.where(weeks >= float(skip_weeks), hA_adj, 0.0)
    hB_eff = np.where(weeks >= float(skip_weeks), hB_adj, 0.0)

    CH_A = np.nancumsum(hA_eff)
    CH_B = np.nancumsum(hB_eff)
    K_raw = np.full_like(CH_A, np.nan, dtype=float)
    # Avoid divide warnings by using numpy's masked divide
    np.divide(CH_A, CH_B, out=K_raw, where=(CH_B > 0))

    # Normalize at effective baseline index
    t0_idx = int(skip_weeks + normalization_weeks - 1)
    t0_idx = min(max(t0_idx, 0), len(K_raw) - 1)
    baseline = K_raw[t0_idx]
    if not (np.isfinite(baseline) and baseline > 0):
        baseline = 1.0

    return np.where(np.isfinite(K_raw), K_raw / baseline, np.nan)


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate pathological negative-control figures for paper.")
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path("example/Frail_cohort_mix.xlsx"),
        help="Path to Frail_cohort_mix.xlsx",
    )
    ap.add_argument(
        "--sheet",
        type=str,
        default="1-10 frailty",
        help="Sheet name containing the pathological frailty mixture",
    )
    ap.add_argument(
        "--outdir",
        type=Path,
        default=Path("documentation/preprint/figs"),
        help="Output directory for PNG figures and JSON parameters",
    )
    ap.add_argument(
        "--skip_weeks",
        type=int,
        default=2,
        help="Number of initial weeks to exclude from cumulative hazard accumulation (matches KCOR.py default)",
    )
    ap.add_argument(
        "--norm_weeks",
        type=int,
        default=4,
        help="Number of weeks of accumulated data used to define the KCOR baseline normalization week (matches KCOR.py default)",
    )
    ap.add_argument(
        "--weights_b",
        type=str,
        default="0.30,0.20,0.20,0.20,0.10",
        help="Comma-separated mixture weights for Cohort B across the same frailty levels (must sum ~1)",
    )
    ap.add_argument(
        "--N0",
        type=float,
        default=1000.0,
        help="Total initial cohort size used for deterministic simulation (scale does not matter)",
    )
    ap.add_argument(
        "--max_week",
        type=int,
        default=None,
        help="Optional maximum week to include (e.g. 250). Default: use full sheet length.",
    )
    args = ap.parse_args()

    data = read_pathological_sheet(args.xlsx, args.sheet)
    weeks = data.weeks
    p_by_group = data.p_by_group

    if args.max_week is not None:
        keep = weeks <= float(args.max_week)
        weeks = weeks[keep]
        p_by_group = p_by_group[keep, :]
        hazard_sheet = data.hazard_sheet[keep]
    else:
        hazard_sheet = data.hazard_sheet

    # Cohort A weights from the sheet (equal by default in the example)
    wA = data.alive0_by_group / float(np.sum(data.alive0_by_group))
    # Cohort B weights (user-configurable)
    wB = np.array([float(x) for x in args.weights_b.split(",")], dtype=float)
    wB = wB / wB.sum()

    # Simulate cohort hazards deterministically using per-group probabilities from the sheet
    hA = simulate_cohort_hazard_from_p(p_by_group=p_by_group, weights=wA, N0=args.N0)
    hB = simulate_cohort_hazard_from_p(p_by_group=p_by_group, weights=wB, N0=args.N0)

    # Fit relaxation model to each cohort hazard
    fitA = fit_relaxation_model(weeks, hA)
    fitB = fit_relaxation_model(weeks, hB)

    # Generate fit overlay (use Cohort A)
    predA = np.exp(relaxation_log_hazard(weeks, C=fitA.C, k0=fitA.k0, k_inf=fitA.k_inf, tau=fitA.tau))

    # KCOR under null
    kcor = compute_kcor(
        weeks,
        hA=hA,
        hB=hB,
        skip_weeks=int(args.skip_weeks),
        normalization_weeks=int(args.norm_weeks),
        fitA=fitA,
        fitB=fitB,
    )

    # Write outputs
    outdir = args.outdir
    _ensure_dir(outdir)

    params = {
        "xlsx": str(args.xlsx),
        "sheet": args.sheet,
        "weekly_log_slope": data.weekly_log_slope,
        "base_prob": data.base_prob,
        "frailty": data.frailty,
        "weights_A": wA.tolist(),
        "weights_B": wB.tolist(),
        "skip_weeks": int(args.skip_weeks),
        "norm_weeks": int(args.norm_weeks),
        "kcor_baseline_index": int(args.skip_weeks + args.norm_weeks - 1),
        "fitA": fitA.__dict__,
        "fitB": fitB.__dict__,
        "fitA_rms_log_error": fitA.rms_log_error,
        "fitB_rms_log_error": fitB.rms_log_error,
    }
    (outdir / "neg_control_pathological_params.json").write_text(json.dumps(params, indent=2), encoding="utf-8")

    # Plotting is optional dependency; import lazily.
    # Force a non-interactive backend so this works in headless CI/WSL without an X server.
    import matplotlib  # noqa: WPS433 (local import is intentional)

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt  # noqa: WPS433 (local import is intentional)

    # Figure 1: hazard + fit overlay (Cohort A)
    fig1 = plt.figure(figsize=(9, 4.8))
    ax = fig1.add_subplot(1, 1, 1)
    ax.plot(weeks, hA, label="Pathological mixture hazard (synthetic)", linewidth=2)
    ax.plot(weeks, predA, label="4-parameter relaxation fit", linestyle="--", linewidth=2)
    ax.set_xlabel("Week since enrollment")
    ax.set_ylabel("Cohort hazard h(t)")
    ax.set_title("Pathological frailty-mixture hazard and 4-parameter fit")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best")
    # Annotate fit quality (log-RMS)
    ax.text(
        0.01,
        0.01,
        f"Fit RMS(log-error) = {fitA.rms_log_error:.3e}",
        transform=ax.transAxes,
        fontsize=9,
        va="bottom",
        ha="left",
    )
    fig1.tight_layout()
    fig1_path = outdir / "neg_control_pathological_fit.png"
    fig1.savefig(fig1_path, dpi=200)
    plt.close(fig1)

    # Figure 2: pre-normalization hazards + KCOR
    fig2 = plt.figure(figsize=(9, 6.5))
    gs = fig2.add_gridspec(2, 1, height_ratios=[1, 1.1], hspace=0.25)

    ax_top = fig2.add_subplot(gs[0, 0])
    ax_top.plot(weeks, hA, label="Cohort A hazard", linewidth=2)
    ax_top.plot(weeks, hB, label="Cohort B hazard (different mixture)", linewidth=2)
    ax_top.set_ylabel("Cohort hazard h(t)")
    ax_top.set_title("Null effect, different selection/mixture → different curvature (pre-normalization)")
    ax_top.grid(True, alpha=0.3)
    ax_top.legend(loc="best")

    ax_bot = fig2.add_subplot(gs[1, 0])
    ax_bot.plot(weeks, kcor, label="KCOR(t) after curvature normalization + anchoring", linewidth=2)
    ax_bot.axhline(1.0, color="black", linestyle="--", linewidth=1, label="KCOR=1 (null)")
    # Visualize skip and baseline week
    baseline_idx = int(args.skip_weeks + args.norm_weeks - 1)
    baseline_week = float(weeks[baseline_idx]) if 0 <= baseline_idx < len(weeks) else float(args.skip_weeks + args.norm_weeks - 1)
    ax_bot.axvspan(0, float(args.skip_weeks), alpha=0.08, label=f"Skip first {args.skip_weeks} weeks")
    ax_bot.axvline(baseline_week, color="gray", linestyle=":", linewidth=1, label=f"Baseline week (t0={baseline_idx})")
    ax_bot.set_xlabel("Week since enrollment")
    ax_bot.set_ylabel("KCOR(t)")
    ax_bot.set_title("Post-normalization: KCOR remains flat at 1 under null")
    ax_bot.grid(True, alpha=0.3)
    ax_bot.legend(loc="best")

    fig2.tight_layout()
    fig2_path = outdir / "neg_control_pathological_kcor.png"
    fig2.savefig(fig2_path, dpi=200)
    plt.close(fig2)

    # Optional: write a quick CSV-like summary to stdout for convenience
    print("Wrote:")
    print(f"  {fig1_path}")
    print(f"  {fig2_path}")
    print(f"  {outdir / 'neg_control_pathological_params.json'}")
    print()
    print("Cohort A fit:", fitA)
    print("Cohort B fit:", fitB)

    # Small consistency check against sheet hazard (Cohort A, equal weights)
    # This is informational only; we don't fail hard.
    if np.isfinite(hazard_sheet).all() and len(hazard_sheet) == len(hA):
        max_abs = float(np.max(np.abs(hazard_sheet - hA)))
        print(f"Max |hazard_sheet - hazard_simulated| = {max_abs:.3e}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())


