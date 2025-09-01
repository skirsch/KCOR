import openpyxl
from openpyxl import load_workbook

# Load the updated workbook
wb = load_workbook('KCORv4_with_formulas.xlsx')
ws = wb['2022-06 booster']

print("Verifying that formulas were correctly applied in columns BH, BI, BJ...")

# Check the specific columns BH, BI, BJ for rows 17-24
for col in [60, 61, 62]:  # BH, BI, BJ
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"\nColumn {col_letter}:")
    for row in range(17, 25):
        cell = ws.cell(row=row, column=col)
        cell_value = cell.value
        if cell_value is not None:
            print(f"Row {row}: Value={cell_value}")

wb.close()
