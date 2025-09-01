import openpyxl
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('KCORv4.xlsx')
ws = wb['2022-06 booster']

print("Verifying column references L, M, N, O...")

# Look at columns L, M, N, O to understand what they represent
for col in [12, 13, 14, 15]:  # L=12, M=13, N=14, O=15
    col_letter = openpyxl.utils.get_column_letter(col)
    print(f"\nColumn {col_letter}:")
    for row in range(1, 15):
        cell = ws.cell(row=row, column=col)
        cell_value = cell.value
        if cell_value is not None:
            print(f"Row {row}: Value={cell_value}")

# Also check what's in row 9 (column labels) to understand the structure
print("\nRow 9 (column labels):")
for col in range(1, 20):
    cell_value = ws.cell(row=9, column=col).value
    if cell_value is not None:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"Column {col_letter}: {cell_value}")

wb.close()
